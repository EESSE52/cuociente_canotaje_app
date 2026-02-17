# seleccionados_pro_plus_ranking_color.py
# Requisitos: PySide6 openpyxl reportlab
# Ejecutar:
#   python3 -m venv .venv
#   source .venv/bin/activate
#   pip install PySide6 openpyxl reportlab
#   python seleccionados_pro_plus_ranking_color.py

import re
import csv
from dataclasses import dataclass
from datetime import datetime
from typing import List, Tuple

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton,
    QTabWidget, QMessageBox, QGroupBox, QFormLayout, QTableWidget, QTableWidgetItem,
    QFileDialog, QSpinBox, QDoubleSpinBox, QHeaderView, QComboBox
)

# Excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# PDF
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas


# -------------------------
# TEMA (COLOR) - QSS
# -------------------------

APP_QSS = """
/* ---- Base ---- */
QWidget {
    background: #0f172a;
    color: #e2e8f0;
    font-size: 13px;
}
QLabel {
    color: #e2e8f0;
}
QGroupBox {
    border: 1px solid #22304a;
    border-radius: 12px;
    margin-top: 12px;
    padding: 10px;
    background: #0b1224;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 12px;
    padding: 0 8px 0 8px;
    color: #93c5fd;
    font-weight: bold;
}

/* ---- Inputs ---- */
QLineEdit, QSpinBox, QDoubleSpinBox, QComboBox {
    background: #0b1224;
    border: 1px solid #22304a;
    border-radius: 10px;
    padding: 8px;
    selection-background-color: #2563eb;
    selection-color: white;
}
QLineEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus, QComboBox:focus {
    border: 1px solid #60a5fa;
}

/* ---- Buttons ---- */
QPushButton {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                               stop:0 #2563eb, stop:1 #1d4ed8);
    border: 0px;
    border-radius: 12px;
    padding: 10px 14px;
    font-weight: bold;
}
QPushButton:hover {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                               stop:0 #3b82f6, stop:1 #2563eb);
}
QPushButton:pressed {
    background: #1e40af;
}
QPushButton:disabled {
    background: #334155;
    color: #94a3b8;
}

/* Botón "peligro" (rojo) */
QPushButton#danger {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                               stop:0 #ef4444, stop:1 #b91c1c);
}
QPushButton#danger:hover { background: #ef4444; }
QPushButton#danger:pressed { background: #991b1b; }

/* ---- Tabs ---- */
QTabWidget::pane {
    border: 1px solid #22304a;
    border-radius: 14px;
    padding: 6px;
    background: #0b1224;
}
QTabBar::tab {
    background: #0b1224;
    border: 1px solid #22304a;
    padding: 10px 14px;
    border-top-left-radius: 12px;
    border-top-right-radius: 12px;
    margin-right: 6px;
    color: #cbd5e1;
}
QTabBar::tab:hover {
    border: 1px solid #60a5fa;
    color: #e2e8f0;
}
QTabBar::tab:selected {
    background: #111c36;
    border: 1px solid #3b82f6;
    color: #93c5fd;
    font-weight: bold;
}

/* ---- Table ---- */
QTableWidget {
    background: #0b1224;
    border: 1px solid #22304a;
    border-radius: 12px;
    gridline-color: #22304a;
    selection-background-color: #1d4ed8;
    selection-color: white;
}
QHeaderView::section {
    background: #111c36;
    color: #93c5fd;
    border: 0px;
    padding: 8px;
    font-weight: bold;
}
QTableCornerButton::section {
    background: #111c36;
    border: 0px;
}

/* ---- Scrollbars ---- */
QScrollBar:vertical {
    background: #0b1224;
    width: 12px;
    margin: 0px;
    border-radius: 6px;
}
QScrollBar::handle:vertical {
    background: #22304a;
    min-height: 30px;
    border-radius: 6px;
}
QScrollBar::handle:vertical:hover {
    background: #334155;
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    height: 0px;
}
"""


# -------------------------
# Tiempos
# -------------------------

_TIME_RE = re.compile(r"^\s*(?:(\d+):)?(\d{1,2})(?:[.,](\d{1,3}))?\s*$")


def parse_time_to_seconds(text: str) -> float:
    m = _TIME_RE.match(text or "")
    if not m:
        raise ValueError("Formato inválido. Usa 1:45.32 o 105.32")

    minutes = int(m.group(1)) if m.group(1) is not None else 0
    seconds = int(m.group(2))
    frac = m.group(3)

    if m.group(1) is not None and seconds >= 60:
        raise ValueError("Segundos deben ser < 60 si usas m:ss")

    frac_s = 0.0
    if frac is not None:
        frac_norm = frac.ljust(3, "0")[:3]
        frac_s = int(frac_norm) / 1000.0

    return minutes * 60 + seconds + frac_s


def format_seconds_to_time(total_seconds: float) -> str:
    if total_seconds < 0:
        sign = "-"
        total_seconds = abs(total_seconds)
    else:
        sign = ""

    minutes = int(total_seconds // 60)
    seconds = total_seconds - minutes * 60
    sec_int = int(seconds)
    millis = int(round((seconds - sec_int) * 1000))

    if millis == 1000:
        millis = 0
        sec_int += 1
        if sec_int == 60:
            sec_int = 0
            minutes += 1

    return f"{sign}{minutes}:{sec_int:02d}.{millis:03d}"


@dataclass
class RowCalc:
    category: str
    disciplina: str
    sexo: str
    name: str
    club: str
    time_text: str
    atleta_s: float
    testigo_s: float
    diff_s: float
    pct_vs: float
    pct_slower: float
    selected: bool


# -------------------------
# Tabla por categoría
# -------------------------

COL_RANK = 0
COL_NAME = 1
COL_CLUB = 2
COL_TIME = 3
COL_DIFF = 4
COL_PCT_VS = 5
COL_PCT_SLOWER = 6
COL_SELECTED = 7

HEADERS = ["Rank", "Nombre", "Club", "Tiempo", "Dif", "% vs", "% + lento", "Sel"]


class CategoryTab(QWidget):
    def __init__(self, category_name: str, disciplina: str, sexo: str):
        super().__init__()
        self.category_name = category_name
        self.disciplina = disciplina
        self.sexo = sexo

        root = QVBoxLayout(self)

        title = QLabel(f"<b>{category_name}</b>")
        root.addWidget(title)

        cfg_box = QGroupBox("Configuración")
        cfg_form = QFormLayout(cfg_box)

        self.testigo = QLineEdit()
        self.testigo.setPlaceholderText("Ej: 3:45.320 o 225.32")
        cfg_form.addRow("Tiempo testigo:", self.testigo)

        self.cutoff = QDoubleSpinBox()
        self.cutoff.setRange(50.0, 200.0)
        self.cutoff.setDecimals(2)
        self.cutoff.setSingleStep(0.25)
        self.cutoff.setValue(105.00)
        cfg_form.addRow("Corte selección (% vs testigo):", self.cutoff)

        self.max_selected = QSpinBox()
        self.max_selected.setRange(1, 999)
        self.max_selected.setValue(999)
        cfg_form.addRow("Máx. seleccionados (opcional):", self.max_selected)

        root.addWidget(cfg_box)

        # Botonera principal
        btn_row1 = QHBoxLayout()
        self.btn_add = QPushButton("Añadir fila")
        self.btn_del = QPushButton("Borrar fila(s)")
        self.btn_del.setObjectName("danger")  # <-- rojo
        self.btn_calc = QPushButton("Calcular + Ranking")

        btn_row1.addWidget(self.btn_add)
        btn_row1.addWidget(self.btn_del)
        btn_row1.addWidget(self.btn_calc)
        btn_row1.addStretch(1)
        root.addLayout(btn_row1)

        # Botonera import/export extra
        btn_row2 = QHBoxLayout()
        self.btn_import_csv = QPushButton("Importar CSV")
        self.btn_import_xlsx = QPushButton("Importar Excel")
        self.btn_paste = QPushButton("Pegar (desde Excel)")
        self.btn_copy = QPushButton("Copiar resultados")
        self.btn_export_xlsx = QPushButton("Exportar Excel")
        self.btn_export_pdf = QPushButton("Exportar PDF")

        btn_row2.addWidget(self.btn_import_csv)
        btn_row2.addWidget(self.btn_import_xlsx)
        btn_row2.addWidget(self.btn_paste)
        btn_row2.addWidget(self.btn_copy)
        btn_row2.addStretch(1)
        btn_row2.addWidget(self.btn_export_xlsx)
        btn_row2.addWidget(self.btn_export_pdf)
        root.addLayout(btn_row2)

        self.table = QTableWidget(0, len(HEADERS))
        self.table.setHorizontalHeaderLabels(HEADERS)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.ExtendedSelection)
        root.addWidget(self.table)

        help_lbl = QLabel(
            "Importación: CSV/Excel con columnas (Nombre, Club, Tiempo) o pegando desde Excel.\n"
            "Pegar: selecciona 3 columnas en Excel (Nombre | Club | Tiempo), copia y luego 'Pegar'."
        )
        help_lbl.setWordWrap(True)
        root.addWidget(help_lbl)

        # Eventos
        self.btn_add.clicked.connect(self.add_row)
        self.btn_del.clicked.connect(self.delete_selected_rows)
        self.btn_calc.clicked.connect(lambda: self.calculate_and_rank(silent=False))
        self.btn_export_xlsx.clicked.connect(self.export_excel)
        self.btn_export_pdf.clicked.connect(self.export_pdf)
        self.btn_import_csv.clicked.connect(self.import_csv)
        self.btn_import_xlsx.clicked.connect(self.import_excel)
        self.btn_paste.clicked.connect(self.paste_from_clipboard)
        self.btn_copy.clicked.connect(self.copy_results_to_clipboard)

        for _ in range(5):
            self.add_row()

    # ---------- UI helpers ----------

    def _make_item(self, text: str, editable: bool = False, align_right: bool = False) -> QTableWidgetItem:
        it = QTableWidgetItem(text)
        flags = it.flags()
        if not editable:
            flags &= ~Qt.ItemIsEditable
        else:
            flags |= Qt.ItemIsEditable
        it.setFlags(flags)
        if align_right:
            it.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
        return it

    def _set_checkbox(self, row: int, checked: bool):
        it = QTableWidgetItem("")
        it.setFlags((it.flags() | Qt.ItemIsUserCheckable) & ~Qt.ItemIsEditable)
        it.setCheckState(Qt.Checked if checked else Qt.Unchecked)
        it.setTextAlignment(Qt.AlignCenter)
        self.table.setItem(row, COL_SELECTED, it)

    def add_row(self, name: str = "", club: str = "", time_text: str = ""):
        r = self.table.rowCount()
        self.table.insertRow(r)

        self.table.setItem(r, COL_RANK, self._make_item("-", editable=False, align_right=True))
        self.table.setItem(r, COL_NAME, self._make_item(name, editable=True))
        self.table.setItem(r, COL_CLUB, self._make_item(club, editable=True))
        self.table.setItem(r, COL_TIME, self._make_item(time_text, editable=True))

        self.table.setItem(r, COL_DIFF, self._make_item("-", editable=False, align_right=True))
        self.table.setItem(r, COL_PCT_VS, self._make_item("-", editable=False, align_right=True))
        self.table.setItem(r, COL_PCT_SLOWER, self._make_item("-", editable=False, align_right=True))
        self._set_checkbox(r, False)

    def delete_selected_rows(self):
        rows = sorted({idx.row() for idx in self.table.selectedIndexes()}, reverse=True)
        if not rows:
            QMessageBox.information(self, "Info", "Selecciona filas para borrar.")
            return
        for r in rows:
            self.table.removeRow(r)

    def _read_row_inputs(self, r: int) -> Tuple[str, str, str, bool]:
        name = (self.table.item(r, COL_NAME).text() or "").strip()
        club = (self.table.item(r, COL_CLUB).text() or "").strip()
        time_text = (self.table.item(r, COL_TIME).text() or "").strip()
        sel_item = self.table.item(r, COL_SELECTED)
        selected = sel_item.checkState() == Qt.Checked if sel_item else False
        return name, club, time_text, selected

    def _write_row_outputs(self, r: int, diff_s: float, pct_vs: float, pct_slower: float):
        self.table.setItem(r, COL_DIFF, self._make_item(format_seconds_to_time(diff_s), editable=False, align_right=True))
        self.table.setItem(r, COL_PCT_VS, self._make_item(f"{pct_vs:.2f} %", editable=False, align_right=True))
        self.table.setItem(r, COL_PCT_SLOWER, self._make_item(f"{pct_slower:+.2f} %", editable=False, align_right=True))

    def _clear_row_outputs(self, r: int):
        self.table.setItem(r, COL_DIFF, self._make_item("-", editable=False, align_right=True))
        self.table.setItem(r, COL_PCT_VS, self._make_item("-", editable=False, align_right=True))
        self.table.setItem(r, COL_PCT_SLOWER, self._make_item("-", editable=False, align_right=True))
        self.table.setItem(r, COL_RANK, self._make_item("-", editable=False, align_right=True))
        self._set_checkbox(r, False)

    # ---------- Cálculo ----------

    def calculate_and_rank(self, silent: bool = False) -> bool:
        try:
            testigo_s = parse_time_to_seconds(self.testigo.text())
            if testigo_s <= 0:
                raise ValueError("Tiempo testigo debe ser > 0")
        except Exception as e:
            if not silent:
                QMessageBox.warning(self, "Error", f"Tiempo testigo inválido: {e}")
            return False

        cutoff = float(self.cutoff.value())
        max_sel = int(self.max_selected.value())

        rows_calc: List[Tuple[int, float]] = []
        for r in range(self.table.rowCount()):
            name, club, time_text, _ = self._read_row_inputs(r)

            if not (name or club or time_text):
                self._clear_row_outputs(r)
                continue
            if not time_text:
                self._clear_row_outputs(r)
                continue

            try:
                atleta_s = parse_time_to_seconds(time_text)
                diff_s = atleta_s - testigo_s
                pct_vs = (atleta_s / testigo_s) * 100.0
                pct_slower = (diff_s / testigo_s) * 100.0
                self._write_row_outputs(r, diff_s, pct_vs, pct_slower)
                rows_calc.append((r, pct_vs))
            except Exception as e:
                self._clear_row_outputs(r)
                if not silent:
                    QMessageBox.warning(self, "Error", f"Fila {r+1}: {e}")
                return False

        rows_calc.sort(key=lambda x: x[1])

        for rank, (r, _) in enumerate(rows_calc, start=1):
            self.table.setItem(r, COL_RANK, self._make_item(str(rank), editable=False, align_right=True))

        selected_count = 0
        for r, pct_vs in rows_calc:
            ok = pct_vs <= cutoff
            if ok and selected_count < max_sel:
                self._set_checkbox(r, True)
                selected_count += 1
            else:
                self._set_checkbox(r, False)

        if not silent:
            QMessageBox.information(self, "OK", f"Listo. Seleccionados: {selected_count}")
        return True

    def collect_results(self) -> List[RowCalc]:
        ok = self.calculate_and_rank(silent=True)
        if not ok:
            return []

        try:
            testigo_s = parse_time_to_seconds(self.testigo.text())
        except Exception:
            return []

        out: List[RowCalc] = []
        for r in range(self.table.rowCount()):
            name, club, time_text, selected = self._read_row_inputs(r)
            if not (name or club or time_text):
                continue
            if not time_text:
                continue

            try:
                atleta_s = parse_time_to_seconds(time_text)
                diff_s = atleta_s - testigo_s
                pct_vs = (atleta_s / testigo_s) * 100.0
                pct_slower = (diff_s / testigo_s) * 100.0
            except Exception:
                continue

            out.append(RowCalc(
                category=self.category_name,
                disciplina=self.disciplina,
                sexo=self.sexo,
                name=name,
                club=club,
                time_text=time_text,
                atleta_s=atleta_s,
                testigo_s=testigo_s,
                diff_s=diff_s,
                pct_vs=pct_vs,
                pct_slower=pct_slower,
                selected=selected
            ))

        out.sort(key=lambda x: x.pct_vs)
        return out

    # ---------- Importar CSV/Excel/Pegar ----------

    def import_csv(self):
        path, _ = QFileDialog.getOpenFileName(self, "Importar CSV", "", "CSV (*.csv)")
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.reader(f)
                rows = list(reader)

            parsed = self._parse_import_rows(rows)
            self._replace_table_with_import(parsed)
            QMessageBox.information(self, "OK", f"Importado CSV: {len(parsed)} filas")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"No se pudo importar CSV:\n{e}")

    def import_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, "Importar Excel", "", "Excel (*.xlsx *.xlsm)")
        if not path:
            return
        try:
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([("" if v is None else str(v)).strip() for v in row])

            parsed = self._parse_import_rows(rows)
            self._replace_table_with_import(parsed)
            QMessageBox.information(self, "OK", f"Importado Excel: {len(parsed)} filas")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"No se pudo importar Excel:\n{e}")

    def paste_from_clipboard(self):
        cb = QApplication.clipboard()
        text = (cb.text() or "").strip()
        if not text:
            QMessageBox.information(self, "Info", "No hay texto en el portapapeles.")
            return
        lines = [ln for ln in text.splitlines() if ln.strip()]
        rows = [ln.split("\t") for ln in lines]

        try:
            parsed = self._parse_import_rows(rows)
            for n, c, t in parsed:
                self.add_row(n, c, t)
            QMessageBox.information(self, "OK", f"Pegado: {len(parsed)} filas")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"No se pudo pegar:\n{e}")

    def _parse_import_rows(self, rows: List[List[str]]) -> List[Tuple[str, str, str]]:
        cleaned = []
        for r in rows:
            r2 = [("" if v is None else str(v)).strip() for v in r]
            if any(x for x in r2):
                cleaned.append(r2)
        if not cleaned:
            return []

        header = [c.lower() for c in cleaned[0]]

        def find_idx(keys):
            for k in keys:
                for i, h in enumerate(header):
                    if k == h or k in h:
                        return i
            return None

        name_idx = find_idx(["nombre", "name"])
        club_idx = find_idx(["club", "equipo", "team"])
        time_idx = find_idx(["tiempo", "time", "marca"])

        start = 1 if (name_idx is not None and time_idx is not None) else 0

        parsed: List[Tuple[str, str, str]] = []
        for r in cleaned[start:]:
            if name_idx is None or time_idx is None:
                n = r[0] if len(r) > 0 else ""
                c = r[1] if len(r) > 1 else ""
                t = r[2] if len(r) > 2 else ""
            else:
                n = r[name_idx] if len(r) > name_idx else ""
                c = r[club_idx] if (club_idx is not None and len(r) > club_idx) else ""
                t = r[time_idx] if len(r) > time_idx else ""

            n, c, t = n.strip(), c.strip(), t.strip()
            if not (n or c or t):
                continue
            parsed.append((n, c, t))

        return parsed

    def _replace_table_with_import(self, parsed: List[Tuple[str, str, str]]):
        self.table.setRowCount(0)
        for n, c, t in parsed:
            self.add_row(n, c, t)

    # ---------- Copiar resultados (categoría) ----------

    def copy_results_to_clipboard(self):
        rows = self.collect_results()
        if not rows:
            QMessageBox.information(self, "Info", "No hay resultados para copiar.")
            return

        cutoff = float(self.cutoff.value())
        max_sel = int(self.max_selected.value())
        testigo_s = rows[0].testigo_s
        now = datetime.now().strftime("%Y-%m-%d %H:%M")

        lines = []
        lines.append("CÁLCULO SELECCIONADOS – CANOTAJE")
        lines.append(f"Categoría: {self.category_name}")
        lines.append(f"Tiempo testigo: {format_seconds_to_time(testigo_s)} | Corte: {cutoff:.2f}% | Máx: {max_sel}")
        lines.append(f"Fecha: {now}")
        lines.append("")
        lines.append("Rank | Nombre | Club | Tiempo | %vs | Dif | Sel")
        lines.append("-" * 70)

        for i, rc in enumerate(rows, start=1):
            sel = "SI" if rc.selected else "NO"
            lines.append(
                f"{i:>4} | {rc.name} | {rc.club} | {rc.time_text} | {rc.pct_vs:6.2f}% | "
                f"{format_seconds_to_time(rc.diff_s):>9} | {sel}"
            )

        QApplication.clipboard().setText("\n".join(lines))
        QMessageBox.information(self, "OK", "Resultados copiados al portapapeles.")

    # ---------- Export Excel/PDF (categoría) ----------

    def export_excel(self):
        rows = self.collect_results()
        if not rows:
            QMessageBox.information(self, "Info", "No hay datos para exportar.")
            return

        suggested = f"seleccionados_{self._safe_filename(self.category_name)}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "Guardar Excel", suggested, "Excel (*.xlsx)")
        if not path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Resultados"

            now = datetime.now().strftime("%Y-%m-%d %H:%M")
            testigo_s = rows[0].testigo_s
            cutoff = float(self.cutoff.value())
            max_sel = int(self.max_selected.value())

            ws["A1"] = "Categoría:"; ws["B1"] = self.category_name
            ws["A2"] = "Tiempo testigo:"; ws["B2"] = format_seconds_to_time(testigo_s)
            ws["A3"] = "Corte (%):"; ws["B3"] = cutoff
            ws["A4"] = "Máx. seleccionados:"; ws["B4"] = max_sel
            ws["A5"] = "Exportado:"; ws["B5"] = now
            for cell in ["A1", "A2", "A3", "A4", "A5"]:
                ws[cell].font = Font(bold=True)

            header_row = 7
            headers = ["Rank", "Nombre", "Club", "Tiempo", "Dif", "% vs Testigo", "% + lento", "Seleccionado"]
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=c, value=h)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for i, rc in enumerate(rows, start=1):
                r = header_row + i
                ws.cell(r, 1, i)
                ws.cell(r, 2, rc.name)
                ws.cell(r, 3, rc.club)
                ws.cell(r, 4, rc.time_text)
                ws.cell(r, 5, format_seconds_to_time(rc.diff_s))
                ws.cell(r, 6, float(f"{rc.pct_vs:.2f}"))
                ws.cell(r, 7, float(f"{rc.pct_slower:.2f}"))
                ws.cell(r, 8, "SI" if rc.selected else "NO")

            widths = [6, 26, 22, 12, 16, 14, 12, 14]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            wb.save(path)
            QMessageBox.information(self, "OK", f"Excel guardado:\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"No se pudo exportar Excel:\n{e}")

    def export_pdf(self):
        rows = self.collect_results()
        if not rows:
            QMessageBox.information(self, "Info", "No hay datos para exportar.")
            return

        suggested = f"seleccionados_{self._safe_filename(self.category_name)}.pdf"
        path, _ = QFileDialog.getSaveFileName(self, "Guardar PDF", suggested, "PDF (*.pdf)")
        if not path:
            return

        try:
            self._write_pdf(path, rows)
            QMessageBox.information(self, "OK", f"PDF guardado:\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"No se pudo exportar PDF:\n{e}")

    def _write_pdf(self, path: str, rows: List[RowCalc]):
        c = canvas.Canvas(path, pagesize=landscape(A4))
        _, H = landscape(A4)

        margin = 12 * mm
        x = margin
        y = H - margin

        title = "Cálculo Seleccionados – % respecto a Tiempo Testigo"
        c.setFont("Helvetica-Bold", 14)
        c.drawString(x, y, title)
        y -= 10 * mm

        c.setFont("Helvetica", 11)
        testigo_s = rows[0].testigo_s
        cutoff = float(self.cutoff.value())
        max_sel = int(self.max_selected.value())
        now = datetime.now().strftime("%Y-%m-%d %H:%M")

        c.drawString(x, y, f"Categoría: {self.category_name}")
        y -= 6 * mm
        c.drawString(x, y, f"Tiempo testigo: {format_seconds_to_time(testigo_s)} | Corte: {cutoff:.2f}% | Máx: {max_sel} | Exportado: {now}")
        y -= 10 * mm

        col_titles = ["Rank", "Nombre", "Club", "Tiempo", "Dif", "% vs", "% + lento", "Sel"]
        col_widths = [14*mm, 55*mm, 45*mm, 22*mm, 28*mm, 18*mm, 22*mm, 12*mm]
        row_h = 7 * mm

        def header():
            nonlocal y
            c.setFont("Helvetica-Bold", 10)
            xx = x
            for t, w in zip(col_titles, col_widths):
                c.drawString(xx + 1.5*mm, y, t)
                xx += w
            y -= row_h
            c.setFont("Helvetica", 10)

        def new_page():
            nonlocal y
            c.showPage()
            y = H - margin
            c.setFont("Helvetica-Bold", 14)
            c.drawString(x, y, title)
            y -= 10 * mm
            c.setFont("Helvetica", 11)
            c.drawString(x, y, f"Categoría: {self.category_name}")
            y -= 6 * mm
            c.drawString(x, y, f"Tiempo testigo: {format_seconds_to_time(testigo_s)} | Corte: {cutoff:.2f}% | Máx: {max_sel} | Exportado: {now}")
            y -= 10 * mm
            header()

        header()
        for i, rc in enumerate(rows, start=1):
            if y < margin + 12*mm:
                new_page()

            vals = [
                str(i), rc.name, rc.club, rc.time_text,
                format_seconds_to_time(rc.diff_s),
                f"{rc.pct_vs:.2f}%",
                f"{rc.pct_slower:+.2f}%",
                "SI" if rc.selected else "NO"
            ]

            xx = x
            for val, w in zip(vals, col_widths):
                txt = self._clip_text(val, w - 3*mm, "Helvetica", 10)
                c.drawString(xx + 1.5*mm, y, txt)
                xx += w
            y -= row_h

        c.save()

    @staticmethod
    def _clip_text(text: str, max_width: float, font: str, size: int) -> str:
        if stringWidth(text, font, size) <= max_width:
            return text
        ell = "..."
        w_ell = stringWidth(ell, font, size)
        if w_ell > max_width:
            return ""
        s = text
        while s and stringWidth(s, font, size) + w_ell > max_width:
            s = s[:-1]
        return s + ell

    @staticmethod
    def _safe_filename(s: str) -> str:
        s = s.lower().strip()
        s = re.sub(r"[^a-z0-9]+", "_", s)
        return s.strip("_")


# -------------------------
# Pestaña Ranking Global
# -------------------------

R_HEADERS = ["Rank", "Disciplina", "Sexo", "Categoría", "Nombre", "Club", "Tiempo", "% vs", "Dif", "Sel"]


class RankingTab(QWidget):
    def __init__(self, category_tabs: List[CategoryTab]):
        super().__init__()
        self.category_tabs = category_tabs

        root = QVBoxLayout(self)

        title = QLabel("<b>RANKING GLOBAL</b> (ordenado por mejor % vs su tiempo testigo)")
        root.addWidget(title)

        controls = QGroupBox("Controles")
        form = QFormLayout(controls)

        self.top_n = QSpinBox()
        self.top_n.setRange(1, 999)
        self.top_n.setValue(3)
        form.addRow("Mostrar Top N:", self.top_n)

        self.filter_disc = QComboBox()
        self.filter_disc.addItems(["Todos", "Kayak", "Canoa"])
        form.addRow("Filtrar disciplina:", self.filter_disc)

        self.filter_sexo = QComboBox()
        self.filter_sexo.addItems(["Todos", "Masculino", "Femenino"])
        form.addRow("Filtrar sexo:", self.filter_sexo)

        root.addWidget(controls)

        btn_row = QHBoxLayout()
        self.btn_refresh = QPushButton("Actualizar Ranking")
        self.btn_copy = QPushButton("Copiar Top N")
        self.btn_export_xlsx = QPushButton("Exportar Excel (Ranking)")
        self.btn_export_pdf = QPushButton("Exportar PDF (Ranking)")
        btn_row.addWidget(self.btn_refresh)
        btn_row.addWidget(self.btn_copy)
        btn_row.addStretch(1)
        btn_row.addWidget(self.btn_export_xlsx)
        btn_row.addWidget(self.btn_export_pdf)
        root.addLayout(btn_row)

        self.table = QTableWidget(0, len(R_HEADERS))
        self.table.setHorizontalHeaderLabels(R_HEADERS)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        root.addWidget(self.table)

        note = QLabel("Nota: para entrar al ranking, cada pestaña debe tener su tiempo testigo y tiempos válidos.")
        note.setWordWrap(True)
        root.addWidget(note)

        self.btn_refresh.clicked.connect(self.refresh)
        self.btn_copy.clicked.connect(self.copy_top_n)
        self.btn_export_xlsx.clicked.connect(self.export_excel)
        self.btn_export_pdf.clicked.connect(self.export_pdf)

    def _make_item(self, text: str, align_right: bool = False) -> QTableWidgetItem:
        it = QTableWidgetItem(text)
        it.setFlags(it.flags() & ~Qt.ItemIsEditable)
        if align_right:
            it.setTextAlignment(Qt.AlignVCenter | Qt.AlignRight)
        return it

    def gather_all(self) -> List[RowCalc]:
        all_rows: List[RowCalc] = []
        for tab in self.category_tabs:
            all_rows.extend(tab.collect_results())

        disc = self.filter_disc.currentText()
        sexo = self.filter_sexo.currentText()

        if disc != "Todos":
            all_rows = [r for r in all_rows if r.disciplina == disc]
        if sexo != "Todos":
            all_rows = [r for r in all_rows if r.sexo == sexo]

        all_rows.sort(key=lambda r: r.pct_vs)
        return all_rows

    def refresh(self):
        rows = self.gather_all()

        self.table.setRowCount(0)
        for i, rc in enumerate(rows, start=1):
            r = self.table.rowCount()
            self.table.insertRow(r)
            self.table.setItem(r, 0, self._make_item(str(i), align_right=True))
            self.table.setItem(r, 1, self._make_item(rc.disciplina))
            self.table.setItem(r, 2, self._make_item(rc.sexo))
            self.table.setItem(r, 3, self._make_item(rc.category))
            self.table.setItem(r, 4, self._make_item(rc.name))
            self.table.setItem(r, 5, self._make_item(rc.club))
            self.table.setItem(r, 6, self._make_item(rc.time_text, align_right=True))
            self.table.setItem(r, 7, self._make_item(f"{rc.pct_vs:.2f} %", align_right=True))
            self.table.setItem(r, 8, self._make_item(format_seconds_to_time(rc.diff_s), align_right=True))
            self.table.setItem(r, 9, self._make_item("SI" if rc.selected else "NO"))

        QMessageBox.information(self, "OK", f"Ranking actualizado: {len(rows)} registros.")

    def top_rows(self) -> List[RowCalc]:
        rows = self.gather_all()
        n = int(self.top_n.value())
        return rows[:n]

    def copy_top_n(self):
        top = self.top_rows()
        if not top:
            QMessageBox.information(self, "Info", "No hay datos para copiar.")
            return

        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        disc = self.filter_disc.currentText()
        sexo = self.filter_sexo.currentText()
        n = int(self.top_n.value())

        lines = []
        lines.append("RANKING GLOBAL – CANOTAJE")
        lines.append(f"Filtro Disciplina: {disc} | Filtro Sexo: {sexo} | Top N: {n}")
        lines.append(f"Fecha: {now}")
        lines.append("")
        lines.append("Rank | Disc | Sexo | Categoría | Nombre | Club | Tiempo | %vs | Dif")
        lines.append("-" * 95)

        for i, rc in enumerate(top, start=1):
            lines.append(
                f"{i:>4} | {rc.disciplina} | {rc.sexo} | {rc.category} | {rc.name} | {rc.club} | "
                f"{rc.time_text} | {rc.pct_vs:6.2f}% | {format_seconds_to_time(rc.diff_s)}"
            )

        QApplication.clipboard().setText("\n".join(lines))
        QMessageBox.information(self, "OK", "Top N copiado al portapapeles.")

    def export_excel(self):
        rows = self.gather_all()
        if not rows:
            QMessageBox.information(self, "Info", "No hay datos para exportar.")
            return

        suggested = "ranking_global.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "Guardar Excel", suggested, "Excel (*.xlsx)")
        if not path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Ranking"

            now = datetime.now().strftime("%Y-%m-%d %H:%M")
            disc = self.filter_disc.currentText()
            sexo = self.filter_sexo.currentText()

            ws["A1"] = "Ranking Global – Canotaje"
            ws["A1"].font = Font(bold=True)
            ws["A2"] = f"Filtro Disciplina: {disc}"
            ws["A3"] = f"Filtro Sexo: {sexo}"
            ws["A4"] = f"Exportado: {now}"

            header_row = 6
            headers = ["Rank", "Disciplina", "Sexo", "Categoría", "Nombre", "Club", "Tiempo", "% vs", "Dif", "Sel"]
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=c, value=h)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for i, rc in enumerate(rows, start=1):
                r = header_row + i
                ws.cell(r, 1, i)
                ws.cell(r, 2, rc.disciplina)
                ws.cell(r, 3, rc.sexo)
                ws.cell(r, 4, rc.category)
                ws.cell(r, 5, rc.name)
                ws.cell(r, 6, rc.club)
                ws.cell(r, 7, rc.time_text)
                ws.cell(r, 8, float(f"{rc.pct_vs:.2f}"))
                ws.cell(r, 9, format_seconds_to_time(rc.diff_s))
                ws.cell(r, 10, "SI" if rc.selected else "NO")

            widths = [6, 12, 12, 26, 26, 22, 12, 10, 12, 8]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            wb.save(path)
            QMessageBox.information(self, "OK", f"Excel guardado:\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"No se pudo exportar Excel:\n{e}")

    def export_pdf(self):
        rows = self.gather_all()
        if not rows:
            QMessageBox.information(self, "Info", "No hay datos para exportar.")
            return

        suggested = "ranking_global.pdf"
        path, _ = QFileDialog.getSaveFileName(self, "Guardar PDF", suggested, "PDF (*.pdf)")
        if not path:
            return

        try:
            self._write_pdf(path, rows)
            QMessageBox.information(self, "OK", f"PDF guardado:\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"No se pudo exportar PDF:\n{e}")

    def _write_pdf(self, path: str, rows: List[RowCalc]):
        c = canvas.Canvas(path, pagesize=landscape(A4))
        _, H = landscape(A4)

        margin = 12 * mm
        x = margin
        y = H - margin

        title = "Ranking Global – % vs Tiempo Testigo"
        c.setFont("Helvetica-Bold", 14)
        c.drawString(x, y, title)
        y -= 10 * mm

        c.setFont("Helvetica", 11)
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        disc = self.filter_disc.currentText()
        sexo = self.filter_sexo.currentText()
        c.drawString(x, y, f"Filtro Disciplina: {disc} | Filtro Sexo: {sexo} | Exportado: {now}")
        y -= 10 * mm

        col_titles = ["Rank", "Disc", "Sexo", "Categoría", "Nombre", "Club", "Tiempo", "%vs", "Dif"]
        col_widths = [14*mm, 18*mm, 22*mm, 52*mm, 45*mm, 40*mm, 22*mm, 18*mm, 22*mm]
        row_h = 7 * mm

        def header():
            nonlocal y
            c.setFont("Helvetica-Bold", 10)
            xx = x
            for t, w in zip(col_titles, col_widths):
                c.drawString(xx + 1.5*mm, y, t)
                xx += w
            y -= row_h
            c.setFont("Helvetica", 10)

        def new_page():
            nonlocal y
            c.showPage()
            y = H - margin
            c.setFont("Helvetica-Bold", 14)
            c.drawString(x, y, title)
            y -= 10 * mm
            c.setFont("Helvetica", 11)
            c.drawString(x, y, f"Filtro Disciplina: {disc} | Filtro Sexo: {sexo} | Exportado: {now}")
            y -= 10 * mm
            header()

        header()

        for i, rc in enumerate(rows, start=1):
            if y < margin + 12*mm:
                new_page()

            vals = [
                str(i),
                rc.disciplina,
                rc.sexo,
                rc.category,
                rc.name,
                rc.club,
                rc.time_text,
                f"{rc.pct_vs:.2f}%",
                format_seconds_to_time(rc.diff_s),
            ]
            xx = x
            for val, w in zip(vals, col_widths):
                txt = self._clip_text(val, w - 3*mm, "Helvetica", 10)
                c.drawString(xx + 1.5*mm, y, txt)
                xx += w
            y -= row_h

        c.save()

    @staticmethod
    def _clip_text(text: str, max_width: float, font: str, size: int) -> str:
        if stringWidth(text, font, size) <= max_width:
            return text
        ell = "..."
        w_ell = stringWidth(ell, font, size)
        if w_ell > max_width:
            return ""
        s = text
        while s and stringWidth(s, font, size) + w_ell > max_width:
            s = s[:-1]
        return s + ell


# -------------------------
# Ventana principal
# -------------------------

class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Seleccionados Canotaje – % respecto a Tiempo Testigo (con Ranking Global + Color)")
        self.resize(1120, 650)

        root = QVBoxLayout(self)

        intro = QLabel(
            "Incluye 4 pestañas (categorías) + pestaña <b>RANKING GLOBAL</b> para ordenar por mejor % vs tiempo testigo."
        )
        intro.setWordWrap(True)
        root.addWidget(intro)

        tabs = QTabWidget()

        tab_km = CategoryTab("Kayak Masculino – 1000 m", "Kayak", "Masculino")
        tab_cm = CategoryTab("Canoa Masculino – 1000 m", "Canoa", "Masculino")
        tab_kf = CategoryTab("Kayak Femenino – 500 m", "Kayak", "Femenino")
        tab_cf = CategoryTab("Canoa Femenina – 200 m", "Canoa", "Femenino")

        tabs.addTab(tab_km, "K1 M 1000")
        tabs.addTab(tab_cm, "C1 M 1000")
        tabs.addTab(tab_kf, "K1 F 500")
        tabs.addTab(tab_cf, "C1 F 200")

        ranking = RankingTab([tab_km, tab_cm, tab_kf, tab_cf])
        tabs.addTab(ranking, "RANKING")

        root.addWidget(tabs)

        foot = QLabel(
            "Uso típico: define tiempo testigo en cada pestaña, ingresa tiempos, calcula. Luego ve a RANKING y presiona 'Actualizar Ranking'."
        )
        foot.setWordWrap(True)
        foot.setStyleSheet("color:#a5b4fc;")  # un toque de color extra
        root.addWidget(foot)


def main():
    app = QApplication([])
    app.setStyleSheet(APP_QSS)  # <-- activa el tema con colores
    w = MainWindow()
    w.show()
    app.exec()


if __name__ == "__main__":
    main()
